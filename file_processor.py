import os
import io
import csv
import time
from typing import Generator, Dict, Any, Optional, List, Tuple, BinaryIO
from concurrent.futures import ThreadPoolExecutor
import streamlit as st

# Try to import optional document processing libraries
# These will be used if available, but the code will still work without them
try:
    from pdfminer.high_level import extract_text_to_fp
    from pdfminer.layout import LAParams
    PDFMINER_AVAILABLE = True
except ImportError:
    PDFMINER_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Default chunk size (1000 characters)
DEFAULT_CHUNK_SIZE = 1000

def extract_text_from_pdf(file_obj: BinaryIO, chunk_size: int = DEFAULT_CHUNK_SIZE) -> Generator[str, None, None]:
    """
    Extract text from a PDF file in chunks
    
    Args:
        file_obj: File-like object containing PDF data
        chunk_size: Size of chunks to yield
        
    Yields:
        Text chunks from the PDF
    """
    if not PDFMINER_AVAILABLE:
        yield "PDF extraction requires pdfminer.six. Please install it with: pip install pdfminer.six"
        return
    
    output_string = io.StringIO()
    
    try:
        # Extract text to string buffer
        extract_text_to_fp(file_obj, output_string, laparams=LAParams())
        text = output_string.getvalue()
        
        # Yield chunks of text
        for i in range(0, len(text), chunk_size):
            yield text[i:i + chunk_size]
    except Exception as e:
        yield f"Error extracting text from PDF: {str(e)}"
    finally:
        output_string.close()

def extract_text_from_docx(file_obj: BinaryIO, chunk_size: int = DEFAULT_CHUNK_SIZE) -> Generator[str, None, None]:
    """
    Extract text from a DOCX file in chunks
    
    Args:
        file_obj: File-like object containing DOCX data
        chunk_size: Size of chunks to yield
        
    Yields:
        Text chunks from the DOCX
    """
    if not DOCX_AVAILABLE:
        yield "DOCX extraction requires python-docx. Please install it with: pip install python-docx"
        return
    
    try:
        # Load the document
        doc = docx.Document(file_obj)
        
        # Extract text
        full_text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        
        # Yield chunks of text
        for i in range(0, len(full_text), chunk_size):
            yield full_text[i:i + chunk_size]
    except Exception as e:
        yield f"Error extracting text from DOCX: {str(e)}"

def extract_text_from_xlsx(file_obj: BinaryIO, chunk_size: int = DEFAULT_CHUNK_SIZE) -> Generator[str, None, None]:
    """
    Extract text from an XLSX file in chunks
    
    Args:
        file_obj: File-like object containing XLSX data
        chunk_size: Size of chunks to yield
        
    Yields:
        Text chunks from the XLSX
    """
    if not OPENPYXL_AVAILABLE:
        yield "XLSX extraction requires openpyxl. Please install it with: pip install openpyxl"
        return
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        
        # Process each sheet
        current_chunk = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if not any(cell for cell in row):
                    continue
                
                # Add row content to current chunk
                row_text = " ".join(str(cell) if cell is not None else "" for cell in row)
                current_chunk += row_text + "\n"
                
                # If chunk is large enough, yield it
                if len(current_chunk) >= chunk_size:
                    yield current_chunk[:chunk_size]
                    current_chunk = current_chunk[chunk_size:]
        
        # Yield any remaining text
        if current_chunk:
            yield current_chunk
    except Exception as e:
        yield f"Error extracting text from XLSX: {str(e)}"

def extract_text_from_csv(file_obj: BinaryIO, chunk_size: int = DEFAULT_CHUNK_SIZE) -> Generator[str, None, None]:
    """
    Extract text from a CSV file in chunks
    
    Args:
        file_obj: File-like object containing CSV data
        chunk_size: Size of chunks to yield
        
    Yields:
        Text chunks from the CSV
    """
    try:
        # Read CSV data and decode to string
        text_io = io.TextIOWrapper(file_obj, encoding='utf-8')
        reader = csv.reader(text_io)
        
        current_chunk = ""
        for row in reader:
            # Skip empty rows
            if not any(cell for cell in row):
                continue
                
            # Add row content to current chunk
            row_text = ",".join(row) + "\n"
            current_chunk += row_text
            
            # If chunk is large enough, yield it
            if len(current_chunk) >= chunk_size:
                yield current_chunk[:chunk_size]
                current_chunk = current_chunk[chunk_size:]
        
        # Yield any remaining text
        if current_chunk:
            yield current_chunk
    except Exception as e:
        yield f"Error extracting text from CSV: {str(e)}"

def extract_text_from_plaintext(file_obj: BinaryIO, chunk_size: int = DEFAULT_CHUNK_SIZE) -> Generator[str, None, None]:
    """
    Extract text from a plaintext file in chunks
    
    Args:
        file_obj: File-like object containing plaintext data
        chunk_size: Size of chunks to yield
        
    Yields:
        Text chunks from the plaintext file
    """
    try:
        # Read text in chunks
        while True:
            chunk = file_obj.read(chunk_size)
            if not chunk:
                break
                
            # Decode bytes to string if needed
            if isinstance(chunk, bytes):
                try:
                    chunk = chunk.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        # Try alternate encoding if UTF-8 fails
                        chunk = chunk.decode('latin-1')
                    except Exception:
                        chunk = f"Error decoding file content"
            
            yield chunk
    except Exception as e:
        yield f"Error reading plaintext file: {str(e)}"

def get_file_extractor(file_type: str):
    """
    Get the appropriate text extractor for a file type
    
    Args:
        file_type: MIME type or file extension
        
    Returns:
        Function to extract text from the file
    """
    # Normalize file type
    file_type = file_type.lower()
    
    # PDF files
    if file_type in ['application/pdf', '.pdf', 'pdf']:
        return extract_text_from_pdf
    
    # Word documents
    if file_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 
                     'application/msword', '.docx', '.doc', 'docx', 'doc']:
        return extract_text_from_docx
    
    # Excel spreadsheets
    if file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     'application/vnd.ms-excel', '.xlsx', '.xls', 'xlsx', 'xls']:
        return extract_text_from_xlsx
    
    # CSV files
    if file_type in ['text/csv', '.csv', 'csv']:
        return extract_text_from_csv
    
    # Default to plaintext for all other types
    return extract_text_from_plaintext

def scan_file_chunks(file_path: str, file_type: str, scanner_func, chunk_size: int = DEFAULT_CHUNK_SIZE, 
                     max_workers: int = 4) -> Tuple[bool, Dict[str, List[str]], float]:
    """
    Scan a file in chunks using parallel processing
    
    Args:
        file_path: Path to the file
        file_type: MIME type or file extension
        scanner_func: Function to scan each chunk (takes a string, returns tuple of (found, patterns))
        chunk_size: Size of each chunk in characters
        max_workers: Maximum number of parallel workers
        
    Returns:
        Tuple containing:
            - Boolean indicating if sensitive information was found
            - Dictionary of all detected patterns
            - Processing time in seconds
    """
    start_time = time.time()
    
    # Get file size to determine best approach
    file_size_bytes = os.path.getsize(file_path)
    file_size_kb = file_size_bytes / 1024
    
    # For small files, we can just read the entire content and use regular scan_text
    # This avoids the overhead of chunking and parallelization for small files
    if file_size_kb < 500:  # 500KB threshold
        try:
            with open(file_path, 'r', errors='ignore') as f:
                content = f.read()
                sensitive_found, detected = scanner_func(content)
                processing_time = time.time() - start_time
                return sensitive_found, detected, processing_time
        except UnicodeDecodeError:
            # If we have trouble reading as text, fall back to binary chunked processing
            pass
    
    # Get the appropriate extractor for this file type
    extractor = get_file_extractor(file_type)
    
    # Dictionary to store all detected patterns
    all_detected = {}
    sensitive_found = False
    
    try:
        with open(file_path, 'rb') as file_obj:
            # For large files, we can optimize by scanning chunks in parallel
            # Extract text in chunks
            chunks = list(extractor(file_obj, chunk_size))
            
            # Adjust worker count based on file size to avoid overhead for smaller files
            if file_size_kb < 1000:  # 1MB
                effective_workers = min(2, max_workers)
            elif file_size_kb < 5000:  # 5MB
                effective_workers = min(3, max_workers)
            else:
                effective_workers = max_workers
            
            # Process chunks in parallel
            with ThreadPoolExecutor(max_workers=effective_workers) as executor:
                # Submit all scan tasks
                future_to_chunk = {executor.submit(scanner_func, chunk): i for i, chunk in enumerate(chunks)}
                
                # Process results as they complete
                for future in future_to_chunk:
                    try:
                        chunk_sensitive, chunk_detected = future.result()
                        
                        # Update overall results
                        sensitive_found = sensitive_found or chunk_sensitive
                        
                        # Merge detected patterns
                        for pattern_name, matches in chunk_detected.items():
                            if pattern_name not in all_detected:
                                all_detected[pattern_name] = []
                            
                            # Add only unique matches to avoid duplicates between chunks
                            for match in matches:
                                if match not in all_detected[pattern_name]:
                                    all_detected[pattern_name].append(match)
                    except Exception as e:
                        print(f"Error processing chunk: {str(e)}")
    except Exception as e:
        print(f"Error processing file: {str(e)}")
    
    processing_time = time.time() - start_time
    return sensitive_found, all_detected, processing_time

# Simplified scanner function for testing
def demo_scan_chunk(text: str) -> Tuple[bool, Dict[str, List[str]]]:
    """
    Simple demonstration scanner that looks for common patterns
    
    Args:
        text: Text chunk to scan
        
    Returns:
        Tuple of (sensitive_found, detected_patterns)
    """
    import re
    
    # Example patterns (simplified)
    patterns = {
        "email": re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'),
        "phone": re.compile(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b'),
        "ssn": re.compile(r'\b\d{3}-\d{2}-\d{4}\b')
    }
    
    detected = {}
    for name, pattern in patterns.items():
        matches = pattern.findall(text)
        if matches:
            detected[name] = matches
    
    return len(detected) > 0, detected

# Usage example:
if __name__ == "__main__":
    # Example: Scan a PDF file
    file_path = "example.pdf"
    if os.path.exists(file_path):
        print(f"Scanning {file_path}...")
        sensitive, patterns, time_taken = scan_file_chunks(file_path, "pdf", demo_scan_chunk)
        print(f"Sensitive info found: {sensitive}")
        print(f"Patterns detected: {patterns}")
        print(f"Processing time: {time_taken:.2f} seconds")
    else:
        print(f"File {file_path} not found.")